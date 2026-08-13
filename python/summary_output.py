import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
EXCEL_DIR = os.path.join(os.path.dirname(__file__), "..", "excel")
OUTPUT_FILE = os.path.join(EXCEL_DIR, "summary_output.xlsx")

def load_clean_data() -> dict:
    us_macro = pd.read_csv(os.path.join(DATA_DIR, "us_macro.csv"), parse_dates=["date"])
    uk_macro = pd.read_csv(os.path.join(DATA_DIR, "uk_macro.csv"), parse_dates=["date"])
    equity_data = pd.read_csv(os.path.join(DATA_DIR, "equity_data.csv"), parse_dates=["date"])

    return {
        "us_macro": us_macro,
        "uk_macro": uk_macro,
        "equity_data": equity_data,
    }

def get_latest_reported(series: pd.Series):
    valid = series.dropna()
    return valid.iloc[-1] if not valid.empty else series.iloc[-1]


def build_snapshot_sheet(wb: Workbook, data: dict) -> None:
    ws = wb.active
    ws.title = "Dashboard Snapshot"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)

    ws["A1"] = "Macro & Equity Dashboard — Current Snapshot"
    ws["A1"].font = title_font
    ws.merge_cells("A1:C1")

    equity_latest = data["equity_data"].iloc[-1]
    snapshot_rows = [
    ("Metric", "US", "UK"),
    ("CPI (YoY %)", round(get_latest_reported(data["us_macro"]["US_CPI"]), 1), round(get_latest_reported(data["uk_macro"]["UK_CPI"]), 1)),
    ("Unemployment (%)", round(get_latest_reported(data["us_macro"]["US_UNEMPLOYMENT"]), 1), round(get_latest_reported(data["uk_macro"]["UK_UNEMPLOYMENT"]), 1)),
    ("GDP Growth (QoQ %)", round(get_latest_reported(data["us_macro"]["US_GDP"]), 1), round(get_latest_reported(data["uk_macro"]["UK_GDP"]), 1)),
]

    start_row = 3
    for i, row in enumerate(snapshot_rows):
        excel_row = start_row + i
        ws.cell(row=excel_row, column=1, value=row[0])
        ws.cell(row=excel_row, column=2, value=row[1])
        ws.cell(row=excel_row, column=3, value=row[2])

    ws.cell(row=start_row, column=1).font = bold
    ws.cell(row=start_row, column=2).font = bold
    ws.cell(row=start_row, column=3).font = bold

    equity_row = start_row + len(snapshot_rows) + 1
    ws.cell(row=equity_row, column=1, value="S&P 500 (latest close)").font = bold
    ws.cell(row=equity_row, column=2, value=round(equity_latest["SP500"], 2))
    ws.cell(row=equity_row + 1, column=1, value="FTSE 100 (latest close)").font = bold
    ws.cell(row=equity_row + 1, column=2, value=round(equity_latest["FTSE100"], 2))

    for col_letter, width in [("A", 25), ("B", 15), ("C", 15)]:
        ws.column_dimensions[col_letter].width = width



def build_data_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(sheet_name)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    date_col_letter = ws.cell(row=1, column=1).column_letter
    for cell in ws[date_col_letter][1:]:
        cell.number_format = "YYYY-MM-DD"

    for i, column in enumerate(df.columns, start=1):
        if column != "date":
            col_letter = ws.cell(row=1, column=i).column_letter
            for cell in ws[col_letter][1:]:
                cell.number_format = "0.0"

    ws.freeze_panes = "A2"

    for i, column in enumerate(df.columns, start=1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = max(12, len(column) + 2)


def main():
    os.makedirs(EXCEL_DIR, exist_ok=True)

    data = load_clean_data()
    wb = Workbook()

    build_snapshot_sheet(wb, data)
    build_data_sheet(wb, "UK Data", data["uk_macro"])
    build_data_sheet(wb, "US Data", data["us_macro"])

    wb.save(OUTPUT_FILE)
    print(f"Saved workbook to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()



